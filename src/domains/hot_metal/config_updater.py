# src/domains/hot_metal/config_updater.py

from datetime import datetime
from openpyxl import load_workbook


class HotMetalConfigUpdater:
    def __init__(self, logger):
        self.logger = logger

    def update_from_excel(self, file_path: str, hm_cfg: dict, run_date: str) -> dict:
        mon = datetime.strptime(run_date, "%d-%b-%Y").strftime("%b").upper()
        yy = datetime.strptime(run_date, "%d-%b-%Y").strftime("%y")

        wb = load_workbook(file_path, read_only=True)
        try:
            target = next(s for s in wb.sheetnames if mon in s.upper() and yy in s)
        except StopIteration:
            raise ValueError(f"No matching HOT METAL sheet for {mon}-{yy}")

        hm = hm_cfg.setdefault("hot_metal_config", {})
        sheets = hm.setdefault("sheets", {})

        if target not in sheets and sheets:
            sheets[target] = sheets.pop(next(iter(sheets)))
            self.logger.warning(f"Reused HOT METAL config block for '{target}'")

        hm["sheet_name"] = target
        self.logger.info(f"HOT_METAL_CONFIG updated → sheet_name: {target}")
        return hm_cfg
