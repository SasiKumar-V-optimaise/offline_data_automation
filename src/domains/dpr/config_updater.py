from dataclasses import dataclass
from datetime import datetime
from calendar import month_abbr
from copy import deepcopy
import re
from typing import Dict, Any, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


_SHEET_RE = re.compile(r"^([A-Za-z]+)\s*'\s*(\d{2})\s*$", re.IGNORECASE)


def parse_sheet_date(name: str) -> Optional[Tuple[int, int]]:
    if not name:
        return None
    name = name.strip()
    m = _SHEET_RE.match(name)
    if not m:
        return None

    try:
        month_str = m.group(1)[:3].title()
        year = 2000 + int(m.group(2))
        month = list(month_abbr).index(month_str)
        return (year, month)
    except Exception:
        return None


def _deep_merge(a: Dict[str, Any], b: Dict[str, Any]) -> Dict[str, Any]:
    out = deepcopy(a or {})
    for k, v in (b or {}).items():
        if isinstance(v, dict) and isinstance(out.get(k), dict):
            out[k] = _deep_merge(out[k], v)
        else:
            out[k] = deepcopy(v)
    return out


@dataclass
class DPRConfigUpdater:
    logger: any

    def _normalize(self, text: str) -> str:
        return re.sub(r"[^a-z0-9]+", " ", str(text).lower()).strip() if text else ""

    def _sheet_key(self, sheet_name: str) -> str:
        return sheet_name.replace("'", "").replace(" ", "")

    def _column_range(self, value) -> range:
        if isinstance(value, str):
            start, end = [part.strip() for part in value.split(":", 1)]
        else:
            start, end = value
        return range(column_index_from_string(start), column_index_from_string(end) + 1)

    def _row_texts(self, ws, label_scan_cols) -> list[tuple[int, set[str]]]:
        col_range = self._column_range(label_scan_cols)
        rows = []

        for row_idx in range(1, ws.max_row + 1):
            texts = {
                self._normalize(ws.cell(row_idx, col).value)
                for col in col_range
                if ws.cell(row_idx, col).value
            }
            if texts:
                rows.append((row_idx, texts))

        return rows

    def _find_alias_row(self, rows: list[tuple[int, set[str]]], aliases: list[str]) -> int:
        normalized_aliases = [self._normalize(alias) for alias in aliases if alias]
        for alias in normalized_aliases:
            for row_idx, texts in rows:
                if alias in texts:
                    return row_idx
        return 0

    def resolve_sheet_config(self, dpr_cfg: Dict[str, Any], target_sheet: str) -> Dict[str, Any]:
        config = dpr_cfg.get("config", {})
        defaults = config.get("defaults", {})
        sheets = config.get("sheets", {})
        sheet_key = self._sheet_key(target_sheet)

        sheet_cfg = sheets.get(sheet_key)
        if sheet_cfg is None:
            sheet_cfg = next(
                (
                    cfg
                    for cfg in sheets.values()
                    if cfg.get("sheet_name", "").strip() == target_sheet.strip()
                ),
                {},
            )

        resolved = _deep_merge(defaults, sheet_cfg)
        resolved["sheet_name"] = target_sheet
        return resolved

    def resolve_rows_from_sheet(self, ws, sheet_cfg: Dict[str, Any]) -> Dict[str, int]:
        row_aliases = sheet_cfg.get("row_aliases", {})
        if row_aliases:
            label_scan_cols = sheet_cfg.get("label_scan_cols", ["A", "G"])
            scanned_rows = self._row_texts(ws, label_scan_cols)
            resolved_rows = {}

            for field_name, aliases in row_aliases.items():
                aliases = aliases if isinstance(aliases, list) else [aliases]
                row_idx = self._find_alias_row(scanned_rows, aliases)
                resolved_rows[field_name] = row_idx
                if not row_idx:
                    self.logger.warning(
                        f"DPR row not found for '{field_name}' using aliases: {aliases}"
                    )

            return resolved_rows

        old_rows = sheet_cfg.get("rows", {})
        if not old_rows:
            return {}

        label_scan_cols = sheet_cfg.get("label_scan_cols", ["A", "G"])
        scanned_rows = self._row_texts(ws, label_scan_cols)
        found_rows: Dict[str, int] = {}

        for label in old_rows:
            row_idx = self._find_alias_row(scanned_rows, [label])
            found_rows[label] = row_idx
            if not row_idx:
                self.logger.warning(f"'{label}' not found in '{sheet_cfg['sheet_name']}'")

        return found_rows

    def select_sheet_for_run_date(self, wb, run_date: str) -> str:
        run_dt = datetime.strptime(run_date, "%d-%b-%Y").date()
        target = (run_dt.year, run_dt.month)

        dated = [(s, parse_sheet_date(s)) for s in wb.sheetnames]
        dated = [(s, d) for s, d in dated if d]

        if not dated:
            raise RuntimeError("No month-named sheets found in DPR workbook.")

        for sname, parsed in dated:
            if parsed == target:
                return sname

        sname, _ = max(dated, key=lambda x: x[1])
        self.logger.warning(f"No matching sheet for {run_date}; using latest: {sname}")
        return sname

    def update_rows_in_config(self, excel_path: str, dpr_cfg: Dict[str, Any], run_date: str) -> Dict[str, Any]:

        wb = load_workbook(excel_path, data_only=True)
        target_sheet = self.select_sheet_for_run_date(wb, run_date)
        ws = wb[target_sheet]

        config = dpr_cfg.setdefault("config", {})
        sheets = config.setdefault("sheets", {})
        sheet_key = self._sheet_key(target_sheet)

        if sheet_key not in sheets:
            sheets[sheet_key] = {"sheet_name": target_sheet}
            self.logger.info(f"DPR month key '{sheet_key}' created from defaults")

        block = self.resolve_sheet_config(dpr_cfg, target_sheet)
        block["sheet_name"] = target_sheet
        block["rows"] = self.resolve_rows_from_sheet(ws, block)
        sheets[sheet_key] = block

        config["sheets"] = sheets
        dpr_cfg["config"] = config

        self.logger.info(f"DPR config updated in-memory for '{target_sheet}'")
        return dpr_cfg
