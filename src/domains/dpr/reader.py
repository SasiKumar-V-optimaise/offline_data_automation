# src/domains/dpr/reader.py

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, Any, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from core.logging import log_file_read
from domains.dpr.config_updater import DPRConfigUpdater


@dataclass
class DPRReader:
    logger: any

    def _normalize(self, text: str) -> str:
        return re.sub(r"[^a-z0-9]+", " ", str(text).lower()).strip() if text else ""

    def _build_reverse_map(self, rename_map) -> dict:
        reverse = {}
        for new_col, labels in (rename_map or {}).items():
            if isinstance(labels, list):
                for label in labels:
                    reverse[self._normalize(label)] = new_col
        return reverse

    def _read_direct_rows(
        self,
        ws,
        rows: dict,
        valid_cols: list[int],
        non_empty_mask: list[bool],
    ) -> dict:
        raw_data = {}
        for field_name, row_idx in (rows or {}).items():
            if not row_idx:
                continue
            values = [ws.cell(row=row_idx, column=col).value for col in valid_cols]
            raw_data[field_name] = [
                value for value, keep in zip(values, non_empty_mask) if keep
            ]
        return raw_data

    def _read_legacy_rows(
        self,
        ws,
        cfg: Dict[str, Any],
        valid_cols: list[int],
        non_empty_mask: list[bool],
    ) -> dict:
        raw_data = {
            label: [ws.cell(row=row_idx, column=col).value for col in valid_cols]
            for label, row_idx in cfg.get("rows", {}).items()
            if row_idx
        }
        reverse_map = self._build_reverse_map(cfg.get("rename_map", {}))
        mapped_data = {}

        for label, values in raw_data.items():
            vals = [v for v, keep in zip(values, non_empty_mask) if keep]
            norm = self._normalize(label)

            if norm in reverse_map:
                col = reverse_map[norm]
            else:
                label_tokens = set(norm.split())
                best_col, best_score = label, 0
                for key, mapped in reverse_map.items():
                    score = len(label_tokens & set(key.split()))
                    if score > best_score:
                        best_score, best_col = score, mapped
                col = best_col if best_score > 0 else label

            mapped_data[col] = vals

        return mapped_data

    def read_for_date(
        self,
        file_path: str,
        dpr_cfg: Dict[str, Any],
        run_date: str,
    ) -> Optional[pd.DataFrame]:
        run_dt = datetime.strptime(run_date, "%d-%b-%Y").date()

        log_file_read(self.logger, file_path, domain="DPR")
        wb = load_workbook(file_path, data_only=True)
        updater = DPRConfigUpdater(self.logger)
        target_sheet = updater.select_sheet_for_run_date(wb, run_date)

        self.logger.info(f"Reading DPR sheet: {target_sheet}")

        if target_sheet not in wb.sheetnames:
            self.logger.warning(f"Sheet '{target_sheet}' not found; skipping")
            return None

        ws = wb[target_sheet]
        cfg = updater.resolve_sheet_config(dpr_cfg, target_sheet)
        if cfg.get("row_aliases") and not cfg.get("rows"):
            cfg["rows"] = updater.resolve_rows_from_sheet(ws, cfg)

        date_row = int(cfg["date_row"])
        col_start, col_end = cfg["date_cols"]

        col_range = range(
            column_index_from_string(col_start),
            column_index_from_string(col_end) + 1,
        )

        parsed_dates, valid_cols = [], []
        for col in col_range:
            val = ws.cell(row=date_row, column=col).value
            parsed = pd.to_datetime(val, errors="coerce")
            if pd.notna(parsed):
                parsed_dates.append(parsed.date())
                valid_cols.append(col)

        row_indices = [row_idx for row_idx in cfg.get("rows", {}).values() if row_idx]
        if not row_indices:
            self.logger.warning(f"No DPR row mapping found for '{target_sheet}'")
            return None

        non_empty_mask = [
            any(
                ws.cell(row=row_idx, column=valid_cols[i]).value is not None
                for row_idx in row_indices
            )
            for i in range(len(valid_cols))
        ]

        filtered_dates = [d for d, keep in zip(parsed_dates, non_empty_mask) if keep]
        df = pd.DataFrame({"Date": filtered_dates})

        if cfg.get("row_aliases"):
            data = self._read_direct_rows(
                ws,
                cfg.get("rows", {}),
                valid_cols,
                non_empty_mask,
            )
        else:
            data = self._read_legacy_rows(ws, cfg, valid_cols, non_empty_mask)

        for col, values in data.items():
            df[col] = values

        df = df[df["Date"] == run_dt].reset_index(drop=True)
        return df if not df.empty else None
