from __future__ import annotations

from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

from core.logging import log_file_read


class FinesAnalysisReader:
    def __init__(self, logger):
        self.logger = logger

    def read(
        self,
        file_path: str,
        sheet_config: dict[str, Any],
    ) -> list[tuple[pd.DataFrame, dict[str, Any], str]]:
        log_file_read(self.logger, file_path, domain="FINES_ANALYSIS")
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        frames: list[tuple[pd.DataFrame, dict[str, Any], str]] = []

        for key, cfg in sheet_config.items():
            sheet = self._resolve_sheet_name(workbook.sheetnames, cfg["sheet_name"])
            if not sheet:
                self.logger.warning(f"Fines analysis sheet missing: {cfg['sheet_name']}")
                continue

            df = self._read_sheet(workbook[sheet], cfg)
            if df.empty:
                self.logger.warning(f"Fines analysis sheet empty: {sheet}")
                continue

            frames.append((df, cfg, sheet))

        workbook.close()
        return frames

    def _read_sheet(self, worksheet, cfg: dict[str, Any]) -> pd.DataFrame:
        min_col, max_col = self._parse_column_range(cfg["columns"])
        header_row = int(cfg["header_row"])
        empty_row_limit = int(cfg.get("empty_row_limit", 50))

        headers = [
            self._header_name(cell.value, idx)
            for idx, cell in enumerate(
                worksheet.iter_rows(
                    min_row=header_row,
                    max_row=header_row,
                    min_col=min_col,
                    max_col=max_col,
                ).__next__(),
                start=min_col,
            )
        ]
        headers = self._dedupe(headers)

        records: list[list[Any]] = []
        empty_rows_seen = 0
        data_seen = False

        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            values = list(row)
            is_empty = all(value is None or str(value).strip() == "" for value in values)

            if is_empty:
                if data_seen:
                    empty_rows_seen += 1
                    if empty_rows_seen >= empty_row_limit:
                        break
                continue

            data_seen = True
            empty_rows_seen = 0
            records.append(values)

        return pd.DataFrame(records, columns=headers).dropna(how="all")

    @staticmethod
    def _parse_column_range(column_range: str) -> tuple[int, int]:
        start, end = [part.strip() for part in column_range.split(":", 1)]
        return column_index_from_string(start), column_index_from_string(end)

    @staticmethod
    def _header_name(value: Any, column_index: int) -> str:
        if value is None or str(value).strip() == "":
            return f"UNNAMED_{column_index}"
        return str(value).strip()

    @staticmethod
    def _dedupe(headers: list[str]) -> list[str]:
        counts: dict[str, int] = {}
        out = []
        for header in headers:
            counts[header] = counts.get(header, 0) + 1
            if counts[header] == 1:
                out.append(header)
            else:
                out.append(f"{header}_{counts[header]}")
        return out

    @staticmethod
    def _normalize_sheet_name(value: str) -> str:
        return " ".join(str(value).lower().split())

    def _resolve_sheet_name(self, sheet_names: list[str], configured_name: str) -> str | None:
        if configured_name in sheet_names:
            return configured_name

        configured = self._normalize_sheet_name(configured_name)
        for sheet_name in sheet_names:
            if self._normalize_sheet_name(sheet_name) == configured:
                return sheet_name

        return None
